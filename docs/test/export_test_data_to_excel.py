"""
📊 EXPORT TEST DATA TO EXCEL
Script để export data từ database ra file Excel, bao gồm:
- Event detections
- User information
- Camera information
- Snapshot images (URLs)
- Timestamps

Author: Healthcare Monitoring System
Date: 12/01/2026
"""

import os
import sys
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import json

# Add parent directories to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Missing required packages. Installing...")
    os.system("pip install pandas openpyxl xlsxwriter")
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv

# Load environment variables
load_dotenv()


class DatabaseExporter:
    """Export test data from database to Excel"""
    
    def __init__(self):
        self.database_url = os.getenv('DATABASE_URL')
        self.conn = None
        
    def connect(self) -> bool:
        """Connect to database"""
        try:
            if not self.database_url:
                print("❌ DATABASE_URL not found in environment")
                return False
                
            self.conn = psycopg2.connect(self.database_url)
            print("✅ Connected to database")
            return True
        except Exception as e:
            print(f"❌ Database connection failed: {e}")
            return False
    
    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()
            print("✅ Database connection closed")
    
    def get_event_detections(self, 
                             start_date: Optional[str] = None, 
                             end_date: Optional[str] = None,
                             limit: int = 10000) -> List[Dict]:
        """Get event detections from database"""
        try:
            with self.conn.cursor(cursor_factory=RealDictCursor) as cursor:
                # Build query
                query = """
                SELECT 
                    ed.event_id,
                    ed.user_id,
                    ed.camera_id,
                    ed.snapshot_id,
                    ed.event_type,
                    ed.event_description,
                    ed.confidence_score,
                    ed.reliability_score,
                    ed.status,
                    ed.lifecycle_state,
                    ed.confirmation_state,
                    ed.verification_status,
                    ed.is_canceled,
                    ed.detected_at,
                    ed.created_at,
                    ed.bounding_boxes,
                    ed.detection_data,
                    ed.context_data,
                    ed.notes,
                    -- User info
                    u.email as user_email,
                    u.full_name as user_full_name,
                    -- Camera info
                    c.camera_name as camera_name,
                    c.location_in_room as camera_location,
                    c.rtsp_url as camera_rtsp_url
                FROM event_detections ed
                LEFT JOIN users u ON ed.user_id = u.user_id
                LEFT JOIN cameras c ON ed.camera_id = c.camera_id
                WHERE 1=1
                """
                
                params = []
                
                if start_date:
                    query += " AND ed.created_at >= %s"
                    params.append(start_date)
                
                if end_date:
                    query += " AND ed.created_at <= %s"
                    params.append(end_date)
                
                query += " ORDER BY ed.created_at DESC LIMIT %s"
                params.append(limit)
                
                cursor.execute(query, params)
                results = cursor.fetchall()
                
                print(f"✅ Found {len(results)} event detections")
                return [dict(row) for row in results]
                
        except Exception as e:
            print(f"❌ Error fetching event detections: {e}")
            return []
    
    def get_snapshots(self, snapshot_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get snapshots for events by snapshot_id"""
        try:
            if not snapshot_ids:
                return {}
            
            # Filter out empty/None values
            valid_ids = [sid for sid in snapshot_ids if sid and sid != 'None']
            if not valid_ids:
                return {}
            
            with self.conn.cursor(cursor_factory=RealDictCursor) as cursor:
                # Get snapshots by snapshot_id - cast to UUID
                query = """
                SELECT 
                    s.snapshot_id,
                    s.camera_id,
                    s.user_id,
                    s.capture_type,
                    s.captured_at,
                    s.cloud_url as snapshot_cloud_url,
                    si.image_id,
                    si.cloud_url as image_url,
                    si.image_path,
                    si.file_size
                FROM snapshots s
                LEFT JOIN snapshot_images si ON s.snapshot_id = si.snapshot_id
                WHERE s.snapshot_id = ANY(%s::uuid[])
                ORDER BY s.captured_at DESC
                """
                
                cursor.execute(query, (valid_ids,))
                results = cursor.fetchall()
                
                # Group by snapshot_id
                snapshots_by_id = {}
                for row in results:
                    snapshot_id = str(row.get('snapshot_id', ''))
                    if snapshot_id not in snapshots_by_id:
                        snapshots_by_id[snapshot_id] = []
                    snapshots_by_id[snapshot_id].append(dict(row))
                
                print(f"✅ Found {len(results)} snapshot images for {len(snapshots_by_id)} snapshots")
                return snapshots_by_id
                
        except Exception as e:
            print(f"❌ Error fetching snapshots: {e}")
            return {}
    
    def get_snapshot_images_by_snapshot_id(self, snapshot_ids: List[str]) -> Dict[str, List[str]]:
        """Get image URLs by snapshot_id"""
        try:
            if not snapshot_ids:
                return {}
            
            # Filter out empty/None values
            valid_ids = [sid for sid in snapshot_ids if sid and sid != 'None']
            if not valid_ids:
                return {}
            
            with self.conn.cursor(cursor_factory=RealDictCursor) as cursor:
                query = """
                SELECT 
                    snapshot_id,
                    cloud_url as image_url
                FROM snapshot_images
                WHERE snapshot_id = ANY(%s::uuid[]) AND cloud_url IS NOT NULL
                ORDER BY created_at DESC
                """
                
                cursor.execute(query, (valid_ids,))
                results = cursor.fetchall()
                
                # Group by snapshot_id
                images_by_snapshot = {}
                for row in results:
                    snapshot_id = str(row.get('snapshot_id', ''))
                    if snapshot_id not in images_by_snapshot:
                        images_by_snapshot[snapshot_id] = []
                    if row.get('image_url'):
                        images_by_snapshot[snapshot_id].append(row['image_url'])
                
                return images_by_snapshot
                
        except Exception as e:
            print(f"❌ Error fetching snapshot images: {e}")
            return {}
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get statistics summary"""
        try:
            with self.conn.cursor(cursor_factory=RealDictCursor) as cursor:
                stats = {}
                
                # Total events by type
                cursor.execute("""
                    SELECT event_type, COUNT(*) as count
                    FROM event_detections
                    GROUP BY event_type
                    ORDER BY count DESC
                """)
                stats['events_by_type'] = [dict(row) for row in cursor.fetchall()]
                
                # Total events by status
                cursor.execute("""
                    SELECT status, COUNT(*) as count
                    FROM event_detections
                    GROUP BY status
                    ORDER BY count DESC
                """)
                stats['events_by_status'] = [dict(row) for row in cursor.fetchall()]
                
                # Events by day (last 30 days)
                cursor.execute("""
                    SELECT 
                        DATE(created_at) as date,
                        COUNT(*) as count
                    FROM event_detections
                    WHERE created_at >= NOW() - INTERVAL '30 days'
                    GROUP BY DATE(created_at)
                    ORDER BY date DESC
                """)
                stats['events_by_day'] = [dict(row) for row in cursor.fetchall()]
                
                # Average confidence by event type
                cursor.execute("""
                    SELECT 
                        event_type,
                        AVG(confidence_score) as avg_confidence,
                        AVG(reliability_score) as avg_reliability,
                        MIN(confidence_score) as min_confidence,
                        MAX(confidence_score) as max_confidence
                    FROM event_detections
                    GROUP BY event_type
                """)
                stats['confidence_by_type'] = [dict(row) for row in cursor.fetchall()]
                
                # Total counts
                cursor.execute("SELECT COUNT(*) as total FROM event_detections")
                stats['total_events'] = cursor.fetchone()['total']
                
                cursor.execute("SELECT COUNT(*) as total FROM users")
                stats['total_users'] = cursor.fetchone()['total']
                
                cursor.execute("SELECT COUNT(*) as total FROM cameras")
                stats['total_cameras'] = cursor.fetchone()['total']
                
                cursor.execute("SELECT COUNT(*) as total FROM snapshots")
                stats['total_snapshots'] = cursor.fetchone()['total']
                
                return stats
                
        except Exception as e:
            print(f"❌ Error fetching statistics: {e}")
            return {}


def export_to_excel(output_path: str, 
                    start_date: Optional[str] = None, 
                    end_date: Optional[str] = None,
                    limit: int = 10000):
    """Export data to Excel file"""
    
    print("=" * 60)
    print("📊 EXPORTING TEST DATA TO EXCEL")
    print("=" * 60)
    
    exporter = DatabaseExporter()
    
    if not exporter.connect():
        return False
    
    try:
        # 1. Get event detections
        print("\n📥 Fetching event detections...")
        events = exporter.get_event_detections(start_date, end_date, limit)
        
        if not events:
            print("⚠️ No events found")
            exporter.close()
            return False
        
        # 2. Get event IDs and snapshot IDs
        event_ids = [str(e['event_id']) for e in events if e.get('event_id')]
        snapshot_ids = [str(e['snapshot_id']) for e in events if e.get('snapshot_id')]
        
        # 3. Get snapshots
        print("\n📥 Fetching snapshots...")
        snapshots_by_id = exporter.get_snapshots(snapshot_ids)
        
        # 4. Get snapshot images
        print("\n📥 Fetching snapshot images...")
        images_by_snapshot = exporter.get_snapshot_images_by_snapshot_id(snapshot_ids)
        
        # 5. Get statistics
        print("\n📥 Fetching statistics...")
        stats = exporter.get_statistics()
        
        # 6. Build DataFrame
        print("\n📊 Building Excel file...")
        
        # Helper function to convert datetime to string
        def format_datetime(dt):
            if dt is None:
                return ''
            if hasattr(dt, 'strftime'):
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            return str(dt)
        
        # Main events data
        events_data = []
        for event in events:
            snapshot_id = str(event.get('snapshot_id', '')) if event.get('snapshot_id') else ''
            
            # Get images for this event
            image_urls = []
            if snapshot_id and snapshot_id in images_by_snapshot:
                image_urls = images_by_snapshot[snapshot_id]
            
            # Also check snapshots_by_id
            if snapshot_id in snapshots_by_id:
                for snap in snapshots_by_id[snapshot_id]:
                    if snap.get('image_url') and snap['image_url'] not in image_urls:
                        image_urls.append(snap['image_url'])
            
            row = {
                'Event ID': str(event.get('event_id', ''))[:8] + '...',
                'Full Event ID': str(event.get('event_id', '')),
                'Event Type': event.get('event_type', ''),
                'Status': event.get('status', ''),
                'Lifecycle State': event.get('lifecycle_state', ''),
                'Confidence Score': round(event.get('confidence_score', 0) or 0, 4),
                'Reliability Score': round(event.get('reliability_score', 0) or 0, 4),
                'Description': (event.get('event_description', '') or '')[:500],
                'Detected At': format_datetime(event.get('detected_at')),
                'Created At': format_datetime(event.get('created_at')),
                'User ID': str(event.get('user_id', ''))[:8] + '...' if event.get('user_id') else '',
                'User Email': event.get('user_email', ''),
                'User Name': event.get('user_full_name', ''),
                'Camera ID': str(event.get('camera_id', ''))[:8] + '...' if event.get('camera_id') else '',
                'Camera Name': event.get('camera_name', ''),
                'Camera Location': event.get('camera_location', ''),
                'Snapshot ID': snapshot_id[:8] + '...' if snapshot_id else '',
                'Image URL 1': image_urls[0] if len(image_urls) > 0 else '',
                'Image URL 2': image_urls[1] if len(image_urls) > 1 else '',
                'Image URL 3': image_urls[2] if len(image_urls) > 2 else '',
                'Image URL 4': image_urls[3] if len(image_urls) > 3 else '',
                'Total Images': len(image_urls),
                'Is Canceled': event.get('is_canceled', False),
                'Verification Status': event.get('verification_status', ''),
            }
            events_data.append(row)
        
        df_events = pd.DataFrame(events_data)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Event Detections
            df_events.to_excel(writer, sheet_name='Event Detections', index=False)
            
            # Sheet 2: Statistics - Events by Type
            if stats.get('events_by_type'):
                df_by_type = pd.DataFrame(stats['events_by_type'])
                df_by_type.to_excel(writer, sheet_name='Events by Type', index=False)
            
            # Sheet 3: Statistics - Events by Status
            if stats.get('events_by_status'):
                df_by_status = pd.DataFrame(stats['events_by_status'])
                df_by_status.to_excel(writer, sheet_name='Events by Status', index=False)
            
            # Sheet 4: Statistics - Events by Day
            if stats.get('events_by_day'):
                df_by_day = pd.DataFrame(stats['events_by_day'])
                df_by_day.to_excel(writer, sheet_name='Events by Day', index=False)
            
            # Sheet 5: Confidence Statistics
            if stats.get('confidence_by_type'):
                df_confidence = pd.DataFrame(stats['confidence_by_type'])
                for col in ['avg_confidence', 'avg_reliability', 'min_confidence', 'max_confidence']:
                    if col in df_confidence.columns:
                        try:
                            df_confidence[col] = pd.to_numeric(df_confidence[col], errors='coerce').round(4)
                        except:
                            pass
                df_confidence.to_excel(writer, sheet_name='Confidence Stats', index=False)
            
            # Sheet 6: Summary
            summary_data = {
                'Metric': [
                    'Total Events',
                    'Total Users',
                    'Total Cameras',
                    'Total Snapshots',
                    'Export Date',
                    'Date Range Start',
                    'Date Range End',
                ],
                'Value': [
                    stats.get('total_events', 0),
                    stats.get('total_users', 0),
                    stats.get('total_cameras', 0),
                    stats.get('total_snapshots', 0),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    start_date or 'All time',
                    end_date or 'All time',
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Style the workbook
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        
        # Style header rows
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Auto-fit column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Style header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(output_path)
        
        print(f"\n✅ Excel file exported successfully!")
        print(f"📁 Output: {output_path}")
        print(f"📊 Total events: {len(events)}")
        print(f"📋 Sheets: {wb.sheetnames}")
        
        exporter.close()
        return True
        
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        exporter.close()
        return False


def main():
    """Main function"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Export test data from database to Excel')
    parser.add_argument('--output', '-o', type=str, 
                        default=f'test_data_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                        help='Output Excel file path')
    parser.add_argument('--start-date', '-s', type=str, 
                        help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end-date', '-e', type=str, 
                        help='End date (YYYY-MM-DD)')
    parser.add_argument('--limit', '-l', type=int, default=10000,
                        help='Maximum number of events to export')
    parser.add_argument('--last-days', '-d', type=int,
                        help='Export data from last N days')
    
    args = parser.parse_args()
    
    # Handle --last-days option
    start_date = args.start_date
    end_date = args.end_date
    
    if args.last_days:
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=args.last_days)).strftime('%Y-%m-%d')
        print(f"📅 Exporting data from last {args.last_days} days: {start_date} to {end_date}")
    
    # Ensure output directory exists
    output_dir = os.path.dirname(args.output)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    success = export_to_excel(
        output_path=args.output,
        start_date=start_date,
        end_date=end_date,
        limit=args.limit
    )
    
    if success:
        print("\n🎉 Export completed successfully!")
    else:
        print("\n❌ Export failed!")
        sys.exit(1)


if __name__ == '__main__':
    main()
